import openpyxl
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


def create_xl(data, proName, name, phonename, n):
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    # 激活工作表
    ws = wb.active
    # 合并单元格
    ws.merge_cells('A1:Z1')
    # 设置第1行的行高为40
    ws.row_dimensions[1].height = 40
    # 初始化序号
    start_number = 1
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=0, max_row=n+2, min_col=0, max_col=14):
        for cell in row:
            cell.border = thin_border

    # 向单元格写入数据
    ws['A1'] = '1.项目名称：'+proName+'\n2.测试人员：'+name
    ws['A1'].alignment = Alignment(wrap_text=True)
    # 定义下拉选项
    options = ["通过", "挂起", "未通过"]
    # 创建数据有效性对象
    dv = DataValidation(type="list", formula1='"{}"'.format(",".join(options)), showDropDown=False)
    dataTop = ['用例编号', '功能模块', '用例名称', '用例详情', '级别', '前置条件', '测试数据', '测试步骤', '预期结果',
            '实际结果', '状态', '测试日期' '测试执行人', '手机号', '审核人']
    for col_num, value in enumerate(dataTop, start=1):
        ws.cell(row=2, column=col_num, value=value)

    # 将数据有效性应用于指定的列（例如A列）
    for row in range(3, n+3):  # 应用到前10行
        cell = ws[f'K{row}']
        ws.add_data_validation(dv)  # 添加数据有效性到工作表
        dv.add(cell)  # 将下拉列表应用到每个单元格

    # 批量添加数据
    for row_num, row_data in enumerate(data, start=3):  # 从第3行开始
        for col_num, value in enumerate(row_data, start=2):
             ws.cell(row=row_num, column=col_num, value=value)  # Excel 列索引从 1 开始

    # 根据B列的内容设置A列的序号
    for row in range(3, n + 3):
        if ws[f'B{row}'].value:  # 如果B列有内容
            ws[f'A{row}'] = start_number
            start_number += 1  # 增加序号
    # ws['B1'] = '功能模块'
    # 保存文件
    testcaseName = proName+'_'+name+'_'+phonename+'.xlsx'
    wb.save(testcaseName)